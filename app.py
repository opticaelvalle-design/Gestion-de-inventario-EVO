from datetime import datetime
import io
import math
import sqlite3
from pathlib import Path

from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook, load_workbook

DB_PATH = Path(__file__).with_name("inventario.db")


app = Flask(__name__)
app.secret_key = "cambia-esta-clave"  # Necesaria para mostrar mensajes flash

# Historial en memoria para permitir deshacer la última lectura de código
lecturas_historial = []
# Registro cronológico de todas las lecturas realizadas
lecturas_registradas = []

storage_locations = []
inventory_items = []
purchase_orders = []
delivery_notes = []
gaveta_asignaciones = {}
gaveta_asignaciones_archivadas = []
gaveta_secuencia = 1
active_delivery_note_id = None

OPTICA_BRANCHES = ["Blanca", "Abarán", "Bajo", "Murcia"]
optica_inventory = {sucursal: [] for sucursal in OPTICA_BRANCHES}


def _registrar_movimiento_optica(producto: dict, sucursal: str, descripcion: str):
    marca_tiempo = datetime.now().strftime("%Y-%m-%d %H:%M")
    producto.setdefault("movimientos", []).append(
        {"fecha": marca_tiempo, "sucursal": sucursal, "descripcion": descripcion}
    )


def _asegurar_sucursal_optica(sucursal: str):
    if sucursal not in optica_inventory:
        optica_inventory[sucursal] = []
    return optica_inventory[sucursal]


def _buscar_producto_optica(sucursal: str, codigo: str):
    inventario = _asegurar_sucursal_optica(sucursal)
    return next(
        (item for item in inventario if item["codigo"].lower() == codigo.lower()), None
    )


def _crear_producto_optica(
    sucursal: str,
    codigo: str,
    nombre: str,
    tipo: str,
    precio_mayor: float,
    precio_pvp: float,
    cantidad: int,
):
    producto = {
        "codigo": codigo,
        "nombre": nombre,
        "tipo": tipo,
        "precio_mayor": precio_mayor,
        "precio_pvp": precio_pvp,
        "cantidad": cantidad,
        "movimientos": [],
    }
    _registrar_movimiento_optica(
        producto,
        sucursal,
        f"Alta inicial con {cantidad} uds en {sucursal}",
    )
    _asegurar_sucursal_optica(sucursal).append(producto)
    return producto


def _traspasar_a_sucursal(origen: str, destino: str, producto: dict, cantidad: int):
    destino_producto = _buscar_producto_optica(destino, producto["codigo"])
    if not destino_producto:
        destino_producto = _crear_producto_optica(
            destino,
            producto["codigo"],
            producto["nombre"],
            producto.get("tipo", ""),
            float(producto.get("precio_mayor", 0)),
            float(producto.get("precio_pvp", 0)),
            0,
        )
    destino_producto["cantidad"] += cantidad
    _registrar_movimiento_optica(
        destino_producto,
        destino,
        f"Recibidas {cantidad} uds desde {origen}",
    )


def _inicializar_optica_demo():
    """Deja los inventarios óptica vacíos para trabajar solo con datos reales."""

    for sucursal in OPTICA_BRANCHES:
        inventario = _asegurar_sucursal_optica(sucursal)
        inventario.clear()


def _importar_excel_optica(archivo, sucursal: str):
    try:
        workbook = load_workbook(archivo, data_only=True)
    except Exception as exc:  # pragma: no cover - validación defensiva
        raise ValueError("No se pudo leer el Excel. Verifica el formato.") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")

    headers = [
        str(cell).strip().lower() if cell is not None else "" for cell in rows[0]
    ]
    header_map = {nombre: idx for idx, nombre in enumerate(headers)}
    required_headers = {"codigo", "nombre", "cantidad"}

    if not required_headers.issubset(header_map):
        raise ValueError(
            "La plantilla debe incluir las columnas: codigo, nombre y cantidad."
        )

    def _leer_valor(row, key, default=None):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return default
        valor = row[idx]
        if valor is None:
            return default
        return valor

    procesadas = creadas = actualizadas = omitidas = 0

    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        procesadas += 1
        codigo = str(_leer_valor(row, "codigo", "")).strip()
        nombre = str(_leer_valor(row, "nombre", "")).strip()
        if not codigo or not nombre:
            omitidas += 1
            continue

        cantidad_valor = _leer_valor(row, "cantidad", 0)
        try:
            cantidad = int(float(cantidad_valor))
        except (TypeError, ValueError):
            omitidas += 1
            continue

        if cantidad < 0:
            omitidas += 1
            continue

        tipo = str(_leer_valor(row, "tipo", "")).strip()
        precio_mayor = _leer_valor(row, "precio_mayor", 0) or 0
        precio_pvp = _leer_valor(row, "precio_pvp", 0) or 0

        try:
            precio_mayor = float(precio_mayor)
        except (TypeError, ValueError):
            precio_mayor = 0.0

        try:
            precio_pvp = float(precio_pvp)
        except (TypeError, ValueError):
            precio_pvp = 0.0

        existente = _buscar_producto_optica(sucursal, codigo)
        if existente:
            existente.update(
                {
                    "nombre": nombre,
                    "tipo": tipo,
                    "precio_mayor": precio_mayor,
                    "precio_pvp": precio_pvp,
                }
            )
            existente["cantidad"] += cantidad
            _registrar_movimiento_optica(
                existente,
                sucursal,
                f"Importación Excel: +{cantidad} uds",
            )
            actualizadas += 1
        else:
            _crear_producto_optica(
                sucursal, codigo, nombre, tipo, precio_mayor, precio_pvp, cantidad
            )
            creadas += 1

    return {
        "procesadas": procesadas,
        "creadas": creadas,
        "actualizadas": actualizadas,
        "omitidas": omitidas,
    }


def _persistir_linea_pedido(pedido_id: int, linea: dict):
    with get_connection() as conn:
        conn.execute(
            """
            UPDATE purchase_order_lines
            SET cantidad_pedida = ?, cantidad_recibida = ?, cantidad_pendiente = ?, descripcion = ?
            WHERE pedido_id = ? AND lower(codigo) = ?
            """,
            (
                linea["cantidad_pedida"],
                linea["cantidad_recibida"],
                linea["cantidad_pendiente"],
                linea["descripcion"],
                pedido_id,
                linea["codigo"].lower(),
            ),
        )


def _insertar_linea_pedido(pedido_id: int, linea: dict):
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO purchase_order_lines (
                pedido_id, codigo, descripcion, cantidad_pedida, cantidad_recibida, cantidad_pendiente
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                pedido_id,
                linea["codigo"],
                linea["descripcion"],
                linea["cantidad_pedida"],
                linea.get("cantidad_recibida", 0),
                linea.get("cantidad_pendiente", linea["cantidad_pedida"]),
            ),
        )


def get_connection():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def _as_datetime(value: str):
    return datetime.fromisoformat(value) if isinstance(value, str) else value


def _init_db_schema():
    with get_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS storage_locations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                tipo TEXT NOT NULL,
                created_at TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'Abierta'
            );

            CREATE TABLE IF NOT EXISTS inventory_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                tipo TEXT DEFAULT '',
                precio_pvo REAL DEFAULT 0,
                precio_pvp REAL DEFAULT 0,
                cantidad INTEGER NOT NULL,
                ubicacion TEXT NOT NULL,
                UNIQUE(codigo, ubicacion)
            );

            CREATE TABLE IF NOT EXISTS purchase_orders (
                id INTEGER PRIMARY KEY,
                nombre TEXT NOT NULL DEFAULT '',
                cliente TEXT NOT NULL,
                fecha TEXT NOT NULL,
                estado TEXT NOT NULL,
                notas TEXT,
                cerrado INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS purchase_order_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pedido_id INTEGER NOT NULL,
                codigo TEXT NOT NULL,
                descripcion TEXT NOT NULL,
                cantidad_pedida INTEGER NOT NULL,
                cantidad_recibida INTEGER NOT NULL,
                cantidad_pendiente INTEGER NOT NULL,
                FOREIGN KEY (pedido_id) REFERENCES purchase_orders(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS delivery_notes (
                id INTEGER PRIMARY KEY,
                numero TEXT NOT NULL,
                fecha TEXT NOT NULL,
                proveedor TEXT NOT NULL,
                fabrica TEXT NOT NULL,
                precio_transporte REAL NOT NULL
            );

            CREATE TABLE IF NOT EXISTS delivery_note_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                albaran_id INTEGER NOT NULL,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                tipo TEXT NOT NULL,
                precio_pvo REAL NOT NULL,
                precio_pvp REAL NOT NULL,
                cantidad INTEGER NOT NULL,
                FOREIGN KEY (albaran_id) REFERENCES delivery_notes(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS gaveta_asignaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pedido_id INTEGER NOT NULL,
                codigo TEXT NOT NULL,
                cliente TEXT NOT NULL,
                descripcion TEXT NOT NULL,
                unidades INTEGER NOT NULL,
                gaveta_nombre TEXT NOT NULL,
                gaveta_tipo TEXT NOT NULL,
                created_at TEXT NOT NULL,
                activa INTEGER NOT NULL DEFAULT 1
            );
            """
        )


def _migrate_inventory_schema():
    with get_connection() as conn:
        schema_row = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'inventory_items'"
        ).fetchone()
        if not schema_row:
            return

        schema_sql = schema_row[0] or ""
        if "codigo TEXT UNIQUE" not in schema_sql:
            return

        conn.executescript(
            """
            ALTER TABLE inventory_items RENAME TO inventory_items_old;
            CREATE TABLE inventory_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                tipo TEXT DEFAULT '',
                precio_pvo REAL DEFAULT 0,
                precio_pvp REAL DEFAULT 0,
                cantidad INTEGER NOT NULL,
                ubicacion TEXT NOT NULL,
                UNIQUE(codigo, ubicacion)
            );
            INSERT INTO inventory_items (codigo, nombre, tipo, precio_pvo, precio_pvp, cantidad, ubicacion)
            SELECT codigo, nombre, '' AS tipo, 0 AS precio_pvo, 0 AS precio_pvp, cantidad, ubicacion FROM inventory_items_old;
            DROP TABLE inventory_items_old;
            """
        )

        existing_columns = {row[1] for row in conn.execute("PRAGMA table_info('inventory_items')")}
        for column, ddl in (
            ("tipo", "ALTER TABLE inventory_items ADD COLUMN tipo TEXT DEFAULT ''"),
            ("precio_pvo", "ALTER TABLE inventory_items ADD COLUMN precio_pvo REAL DEFAULT 0"),
            ("precio_pvp", "ALTER TABLE inventory_items ADD COLUMN precio_pvp REAL DEFAULT 0"),
        ):
            if column not in existing_columns:
                conn.execute(ddl)


def _migrate_purchase_orders_schema():
    with get_connection() as conn:
        columns = {row[1] for row in conn.execute("PRAGMA table_info('purchase_orders')")}
        if not columns:
            return

        if "nombre" not in columns:
            conn.execute("ALTER TABLE purchase_orders ADD COLUMN nombre TEXT DEFAULT ''")

        if "cerrado" not in columns:
            conn.execute(
                "ALTER TABLE purchase_orders ADD COLUMN cerrado INTEGER NOT NULL DEFAULT 0"
            )


def _migrate_storage_locations_schema():
    with get_connection() as conn:
        columns = {row[1] for row in conn.execute("PRAGMA table_info('storage_locations')")}
        if not columns:
            return

        if "estado" not in columns:
            conn.execute(
                "ALTER TABLE storage_locations ADD COLUMN estado TEXT NOT NULL DEFAULT 'Abierta'"
            )


def _migrate_gavetas_schema():
    with get_connection() as conn:
        columns = {row[1] for row in conn.execute("PRAGMA table_info('gaveta_asignaciones')")}
        if not columns:
            return

        if "activa" not in columns:
            conn.execute(
                "ALTER TABLE gaveta_asignaciones ADD COLUMN activa INTEGER NOT NULL DEFAULT 1"
            )


def _seed_if_empty():
    """Inicializa tablas sin insertar datos de demostración."""

    # Las tablas se crean vacías; la aplicación opera exclusivamente con datos reales.


def _load_data():
    global storage_locations, inventory_items, purchase_orders, delivery_notes, gaveta_secuencia, gaveta_asignaciones
    with get_connection() as conn:
        storage_locations = [
            {
                "nombre": row["nombre"],
                "tipo": row["tipo"],
                "estado": row["estado"] if row["estado"] else "Abierta",
                "created_at": _as_datetime(row["created_at"]),
            }
            for row in conn.execute(
                "SELECT nombre, tipo, created_at, estado FROM storage_locations ORDER BY created_at"
            )
        ]

        inventory_items = [
            {
                "id": row["id"],
                "codigo": row["codigo"],
                "nombre": row["nombre"],
                "tipo": row["tipo"] or "",
                "precio_pvo": row["precio_pvo"] if row["precio_pvo"] is not None else 0.0,
                "precio_pvp": row["precio_pvp"] if row["precio_pvp"] is not None else 0.0,
                "cantidad": row["cantidad"],
                "ubicacion": row["ubicacion"],
            }
            for row in conn.execute(
                """
                SELECT id, codigo, nombre, tipo, precio_pvo, precio_pvp, cantidad, ubicacion
                FROM inventory_items ORDER BY codigo
                """
            )
        ]

        purchase_orders = []
        pedidos_rows = conn.execute(
            "SELECT id, nombre, cliente, fecha, estado, notas, cerrado FROM purchase_orders ORDER BY fecha"
        ).fetchall()
        for pedido in pedidos_rows:
            lineas = [
                {
                    "codigo": linea["codigo"],
                    "descripcion": linea["descripcion"],
                    "cantidad_pedida": linea["cantidad_pedida"],
                    "cantidad_recibida": linea["cantidad_recibida"],
                    "cantidad_pendiente": linea["cantidad_pendiente"],
                }
                for linea in conn.execute(
                    """
                    SELECT codigo, descripcion, cantidad_pedida, cantidad_recibida, cantidad_pendiente
                    FROM purchase_order_lines WHERE pedido_id = ? ORDER BY id
                    """,
                    (pedido["id"],),
                )
            ]
            purchase_orders.append(
                {
                    "id": pedido["id"],
                    "nombre": pedido["nombre"],
                    "cliente": pedido["cliente"],
                    "fecha": _as_datetime(pedido["fecha"]),
                    "estado": pedido["estado"],
                    "cerrado": bool(pedido["cerrado"]),
                    "notas": pedido["notas"],
                    "lineas": lineas,
                }
            )

        delivery_notes = []
        albaranes_rows = conn.execute(
            "SELECT id, numero, fecha, proveedor, fabrica, precio_transporte FROM delivery_notes ORDER BY fecha"
        ).fetchall()
        for albaran in albaranes_rows:
            lineas = [
                {
                    "codigo": linea["codigo"],
                    "nombre": linea["nombre"],
                    "tipo": linea["tipo"],
                    "precio_pvo": linea["precio_pvo"],
                    "precio_pvp": linea["precio_pvp"],
                    "cantidad": linea["cantidad"],
                }
                for linea in conn.execute(
                    """
                    SELECT codigo, nombre, tipo, precio_pvo, precio_pvp, cantidad
                    FROM delivery_note_lines WHERE albaran_id = ? ORDER BY id
                    """,
                    (albaran["id"],),
                )
            ]
            delivery_notes.append(
                {
                    "id": albaran["id"],
                    "numero": albaran["numero"],
                    "fecha": _as_datetime(albaran["fecha"]),
                    "proveedor": albaran["proveedor"],
                    "fabrica": albaran["fabrica"],
                    "precio_transporte": albaran["precio_transporte"],
                    "lineas": lineas,
                }
            )

        asignaciones_rows = conn.execute(
            """
            SELECT pedido_id, codigo, cliente, descripcion, unidades, gaveta_nombre, gaveta_tipo, created_at, activa
            FROM gaveta_asignaciones
            """
        ).fetchall()
        gaveta_asignaciones = {}
        gaveta_asignaciones_archivadas.clear()
        for row in asignaciones_rows:
            asignacion = {
                "pedido_id": row["pedido_id"],
                "cliente": row["cliente"],
                "codigo": row["codigo"],
                "descripcion": row["descripcion"],
                "unidades": row["unidades"],
                "gaveta": {
                    "nombre": row["gaveta_nombre"],
                    "tipo": row["gaveta_tipo"],
                    "created_at": _as_datetime(row["created_at"]),
                },
            }
            if row["activa"]:
                gaveta_asignaciones[(row["pedido_id"], row["codigo"].lower())] = asignacion
            else:
                gaveta_asignaciones_archivadas.append(asignacion)
        gaveta_secuencia = len(storage_locations) + 1


def ensure_database():
    _init_db_schema()
    _migrate_inventory_schema()
    _migrate_purchase_orders_schema()
    _migrate_storage_locations_schema()
    _migrate_gavetas_schema()
    _seed_if_empty()
    _load_data()


ensure_database()
_inicializar_optica_demo()


def _articulos_por_codigo(codigo: str):
    return [
        item for item in inventory_items if item["codigo"].lower() == codigo.lower()
    ]


def _resumen_inventario():
    resumen = {}
    for item in inventory_items:
        clave = item["codigo"].lower()
        if clave not in resumen:
            resumen[clave] = {
                "codigo": item["codigo"],
                "nombre": item["nombre"],
                "total_cantidad": 0,
                "ubicaciones": [],
            }

        resumen[clave]["total_cantidad"] += item["cantidad"]
        resumen[clave]["ubicaciones"].append(
            {"ubicacion": item["ubicacion"], "cantidad": item["cantidad"]}
        )

    return sorted(resumen.values(), key=lambda entry: entry["codigo"].lower())


@app.route("/stock-opticas", methods=["GET", "POST"])
def stock_opticas():
    sucursal = (
        request.form.get("sucursal_actual")
        or request.args.get("sucursal")
        or OPTICA_BRANCHES[0]
    )
    if sucursal not in OPTICA_BRANCHES:
        sucursal = OPTICA_BRANCHES[0]

    termino_busqueda = request.args.get("buscar", "").strip().lower()

    if request.method == "POST":
        accion = request.form.get("accion")
        if accion == "nuevo_producto":
            codigo = request.form.get("codigo", "").strip()
            nombre = request.form.get("nombre", "").strip()
            tipo = request.form.get("tipo", "").strip()
            precio_mayor = request.form.get("precio_mayor", type=float, default=0.0)
            precio_pvp = request.form.get("precio_pvp", type=float, default=0.0)
            cantidad = request.form.get("cantidad", type=int, default=0)

            if not codigo or not nombre or cantidad <= 0:
                flash("Introduce código, nombre y una cantidad válida.", "error")
            else:
                existente = _buscar_producto_optica(sucursal, codigo)
                if existente:
                    existente.update(
                        {
                            "nombre": nombre,
                            "tipo": tipo,
                            "precio_mayor": precio_mayor,
                            "precio_pvp": precio_pvp,
                        }
                    )
                    existente["cantidad"] += cantidad
                    _registrar_movimiento_optica(
                        existente,
                        sucursal,
                        f"Actualizado y añadidas {cantidad} uds",
                    )
                else:
                    _crear_producto_optica(
                        sucursal, codigo, nombre, tipo, precio_mayor, precio_pvp, cantidad
                    )
                flash("Producto guardado en el stock de ópticas.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "importar_excel":
            archivo = request.files.get("archivo_excel")
            if not archivo or archivo.filename == "":
                flash("Selecciona un archivo Excel para importar.", "error")
            elif not archivo.filename.lower().endswith(".xlsx"):
                flash("El archivo debe tener extensión .xlsx.", "error")
            else:
                try:
                    resumen = _importar_excel_optica(archivo, sucursal)
                except ValueError as exc:
                    flash(str(exc), "error")
                else:
                    flash(
                        "Importación completada: "
                        f"{resumen['procesadas']} filas procesadas, "
                        f"{resumen['creadas']} creadas, "
                        f"{resumen['actualizadas']} actualizadas, "
                        f"{resumen['omitidas']} omitidas.",
                        "success",
                    )
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "ajustar":
            codigo = request.form.get("codigo", "").strip()
            cantidad = request.form.get("cantidad", type=int, default=0)
            producto = _buscar_producto_optica(sucursal, codigo)
            if not producto:
                flash("No se encontró el producto en la sucursal.", "error")
            elif cantidad <= 0:
                flash("Indica unidades válidas para añadir.", "warning")
            else:
                producto["cantidad"] += cantidad
                _registrar_movimiento_optica(
                    producto,
                    sucursal,
                    f"Añadidas {cantidad} uds",
                )
                flash("Stock actualizado.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "retirar":
            codigo = request.form.get("codigo", "").strip()
            cantidad = request.form.get("cantidad", type=int, default=0)
            motivo = request.form.get("motivo", "").strip() or "Sin motivo"
            producto = _buscar_producto_optica(sucursal, codigo)
            if not producto:
                flash("No se encontró el producto en la sucursal.", "error")
            elif cantidad <= 0:
                flash("Indica unidades válidas para retirar.", "warning")
            elif producto["cantidad"] < cantidad:
                flash("No hay suficientes unidades para retirar.", "error")
            else:
                producto["cantidad"] -= cantidad
                _registrar_movimiento_optica(
                    producto,
                    sucursal,
                    f"Retiradas {cantidad} uds. Motivo: {motivo}",
                )
                flash("Retirada registrada.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "transferir":
            codigo = request.form.get("codigo", "").strip()
            destino = request.form.get("destino")
            cantidad = request.form.get("cantidad", type=int, default=0)
            producto = _buscar_producto_optica(sucursal, codigo)
            if destino not in OPTICA_BRANCHES:
                flash("Selecciona una sucursal destino válida.", "warning")
            elif destino == sucursal:
                flash("Elige una sucursal distinta para transferir.", "warning")
            elif not producto:
                flash("No se encontró el producto en la sucursal.", "error")
            elif cantidad <= 0:
                flash("Indica unidades válidas para transferir.", "warning")
            elif producto["cantidad"] < cantidad:
                flash("No hay suficientes unidades para transferir.", "error")
            else:
                producto["cantidad"] -= cantidad
                _registrar_movimiento_optica(
                    producto,
                    sucursal,
                    f"Transferidas {cantidad} uds a {destino}",
                )
                _traspasar_a_sucursal(sucursal, destino, producto, cantidad)
                flash("Transferencia completada.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "actualizar_producto":
            codigo_original = request.form.get("codigo_original", "").strip()
            producto = _buscar_producto_optica(sucursal, codigo_original)
            if not producto:
                flash("No se encontró el producto seleccionado.", "error")
                return redirect(url_for("stock_opticas", sucursal=sucursal))

            codigo = request.form.get("codigo", "").strip()
            nombre = request.form.get("nombre", "").strip()
            tipo = request.form.get("tipo", "").strip()
            precio_mayor = request.form.get("precio_mayor", type=float, default=0.0)
            precio_pvp = request.form.get("precio_pvp", type=float, default=0.0)
            cantidad = request.form.get("cantidad", type=int, default=producto["cantidad"])

            if not codigo or not nombre or cantidad < 0:
                flash("Introduce código, nombre y una cantidad válida.", "error")
                return redirect(url_for("stock_opticas", sucursal=sucursal))

            existente = _buscar_producto_optica(sucursal, codigo)
            if existente and existente is not producto:
                flash("Ya existe otro producto con ese código en la sucursal.", "warning")
                return redirect(url_for("stock_opticas", sucursal=sucursal))

            cambios = []
            if producto["codigo"] != codigo:
                cambios.append(f"Código: {producto['codigo']} → {codigo}")
            if producto["nombre"] != nombre:
                cambios.append(f"Nombre: {producto['nombre']} → {nombre}")
            if producto.get("tipo", "") != tipo:
                cambios.append(f"Tipo: {producto.get('tipo', '')} → {tipo}")
            if float(producto.get("precio_mayor", 0)) != float(precio_mayor):
                cambios.append(
                    f"Precio mayorista: {producto.get('precio_mayor', 0)} → {precio_mayor}"
                )
            if float(producto.get("precio_pvp", 0)) != float(precio_pvp):
                cambios.append(
                    f"Precio PVP: {producto.get('precio_pvp', 0)} → {precio_pvp}"
                )
            if int(producto.get("cantidad", 0)) != int(cantidad):
                cambios.append(
                    f"Cantidad: {producto.get('cantidad', 0)} → {cantidad}"
                )

            producto.update(
                {
                    "codigo": codigo,
                    "nombre": nombre,
                    "tipo": tipo,
                    "precio_mayor": precio_mayor,
                    "precio_pvp": precio_pvp,
                    "cantidad": cantidad,
                }
            )

            descripcion_cambios = ", ".join(cambios) if cambios else "Sin cambios"
            _registrar_movimiento_optica(
                producto,
                sucursal,
                f"Actualización de detalle: {descripcion_cambios}",
            )
            flash("Producto actualizado correctamente.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "eliminar_producto":
            codigo = request.form.get("codigo", "").strip()
            producto = _buscar_producto_optica(sucursal, codigo)
            if not producto:
                flash("No se encontró el producto a eliminar.", "error")
            else:
                inventario = _asegurar_sucursal_optica(sucursal)
                inventario.remove(producto)
                flash(f"Producto {codigo} eliminado de {sucursal}.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

        if accion == "lectura_codigo":
            codigo = request.form.get("codigo_barras", "").strip()
            if not codigo:
                flash("Escanea un código válido.", "warning")
            else:
                producto = _buscar_producto_optica(sucursal, codigo)
                if not producto:
                    producto = _crear_producto_optica(
                        sucursal,
                        codigo,
                        f"Artículo {codigo}",
                        "Código de barras",
                        0.0,
                        0.0,
                        0,
                    )
                producto["cantidad"] += 1
                _registrar_movimiento_optica(
                    producto,
                    sucursal,
                    "Entrada por lectura de código de barras",
                )
                flash("Artículo registrado por código de barras.", "success")
            return redirect(url_for("stock_opticas", sucursal=sucursal))

    productos = sorted(
        [
            item
            for item in _asegurar_sucursal_optica(sucursal)
            if not termino_busqueda
            or termino_busqueda in item["codigo"].lower()
            or termino_busqueda in item["nombre"].lower()
        ],
        key=lambda item: item["nombre"].lower(),
    )
    totales_sucursal = sum(
        item["cantidad"] for item in _asegurar_sucursal_optica(sucursal)
    )
    return render_template(
        "stock_opticas.html",
        sucursal=sucursal,
        sucursales=OPTICA_BRANCHES,
        productos=productos,
        totales_sucursal=totales_sucursal,
        termino_busqueda=termino_busqueda,
    )


@app.route("/stock-opticas/plantilla")
def descargar_plantilla_stock_opticas():
    headers = ["codigo", "nombre", "tipo", "precio_mayor", "precio_pvp", "cantidad"]
    output = _crear_excel(headers, [], "Plantilla stock ópticas")
    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_stock_opticas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/stock-opticas/exportar")
def exportar_stock_opticas():
    sucursal = request.args.get("sucursal") or OPTICA_BRANCHES[0]
    if sucursal not in OPTICA_BRANCHES:
        sucursal = OPTICA_BRANCHES[0]

    headers = ["Código", "Nombre", "Tipo", "Precio mayorista", "Precio PVP", "Cantidad"]
    productos = _asegurar_sucursal_optica(sucursal)
    rows = [
        [
            item["codigo"],
            item["nombre"],
            item.get("tipo", ""),
            item.get("precio_mayor", 0),
            item.get("precio_pvp", 0),
            item.get("cantidad", 0),
        ]
        for item in productos
    ]

    output = _crear_excel(headers, rows, f"Stock {sucursal}")
    filename = (
        f"stock_opticas_{sucursal.replace(' ', '_').lower()}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/crear-gavetas", methods=["GET", "POST"])
def crear_gavetas():
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        tipo = request.form.get("tipo", "").strip()

        if not nombre or not tipo:
            flash("El nombre y el tipo de la ubicación son obligatorios.", "error")
        else:
            nueva_ubicacion = {
                "nombre": nombre,
                "tipo": tipo,
                "estado": "Abierta",
                "created_at": datetime.now(),
            }
            storage_locations.append(nueva_ubicacion)
            with get_connection() as conn:
                conn.execute(
                    "INSERT INTO storage_locations (nombre, tipo, created_at, estado) VALUES (?, ?, ?, ?)",
                    (
                        nueva_ubicacion["nombre"],
                        nueva_ubicacion["tipo"],
                        nueva_ubicacion["created_at"].isoformat(),
                        nueva_ubicacion["estado"],
                    ),
                )
            flash("Ubicación registrada correctamente.", "success")
        return redirect(url_for("crear_gavetas"))

    return render_template("crear_gavetas.html", ubicaciones=storage_locations)


@app.route("/crear-gavetas/exportar")
def exportar_gavetas():
    headers = [
        "Nombre",
        "Tipo",
        "Fecha de alta",
        "Unidades en inventario",
        "Unidades asignadas",
        "Unidades totales",
    ]

    rows = []
    for ubicacion in storage_locations:
        unidades_inventario = sum(
            item["cantidad"] for item in inventory_items if item["ubicacion"].lower() == ubicacion["nombre"].lower()
        )
        unidades_asignadas = sum(
            asignacion["unidades"]
            for asignacion in gaveta_asignaciones.values()
            if asignacion["gaveta"]["nombre"].lower() == ubicacion["nombre"].lower()
        )
        rows.append(
            [
                ubicacion["nombre"],
                ubicacion["tipo"],
                _as_datetime(ubicacion["created_at"]).strftime("%Y-%m-%d %H:%M"),
                unidades_inventario,
                unidades_asignadas,
                unidades_inventario + unidades_asignadas,
            ]
        )

    output = _crear_excel(headers, rows, "Gavetas")
    filename = f"gavetas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/crear-gavetas/<path:nombre>/exportar-csv")
def exportar_gaveta_csv(nombre: str):
    ubicacion = next(
        (
            ubicacion
            for ubicacion in storage_locations
            if ubicacion["nombre"].lower() == nombre.lower()
        ),
        None,
    )
    if not ubicacion:
        flash("No se encontró la gaveta solicitada.", "error")
        return redirect(url_for("crear_gavetas"))

    recuento_productos = {}

    for articulo in inventory_items:
        if articulo["ubicacion"].lower() != ubicacion["nombre"].lower():
            continue

        producto = recuento_productos.setdefault(
            articulo["codigo"],
            {
                "codigo": articulo["codigo"],
                "descripcion": articulo.get("nombre", ""),
                "inventario": 0,
                "asignadas": 0,
            },
        )
        producto["inventario"] += articulo.get("cantidad", 0)
        if not producto["descripcion"]:
            producto["descripcion"] = articulo.get("nombre", "")

    for asignacion in gaveta_asignaciones.values():
        if asignacion["gaveta"]["nombre"].lower() != ubicacion["nombre"].lower():
            continue

        producto = recuento_productos.setdefault(
            asignacion["codigo"],
            {
                "codigo": asignacion["codigo"],
                "descripcion": asignacion.get("descripcion", ""),
                "inventario": 0,
                "asignadas": 0,
            },
        )
        producto["asignadas"] += asignacion.get("unidades", 0)
        if not producto["descripcion"]:
            producto["descripcion"] = asignacion.get("descripcion", "")

    headers = [
        "Código",
        "Descripción",
        "Unidades en inventario",
        "Unidades asignadas",
        "Total en gaveta",
    ]

    rows = []
    for producto in sorted(recuento_productos.values(), key=lambda item: item["codigo"].lower()):
        total = producto["inventario"] + producto["asignadas"]
        rows.append(
            [
                producto["codigo"],
                producto["descripcion"],
                producto["inventario"],
                producto["asignadas"],
                total,
            ]
        )

    output = _crear_excel(
        headers,
        rows,
        f"Gaveta {_sanitize_sheet_name(ubicacion['nombre'])[:31] or 'Gaveta'}",
    )
    filename = f"gaveta_{ubicacion['nombre'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/crear-gavetas/<path:nombre>/renombrar", methods=["POST"])
def renombrar_gaveta(nombre: str):
    global storage_locations, inventory_items, gaveta_asignaciones

    ubicacion = next(
        (
            ubicacion
            for ubicacion in storage_locations
            if ubicacion["nombre"].lower() == nombre.lower()
        ),
        None,
    )
    if not ubicacion:
        flash("No se encontró la gaveta a renombrar.", "error")
        return redirect(url_for("crear_gavetas"))

    nuevo_nombre = request.form.get("nuevo_nombre", "").strip()
    if not nuevo_nombre:
        flash("Introduce un nuevo nombre para la gaveta.", "error")
        return redirect(url_for("gaveta_detalle", nombre=nombre))

    if any(ubic["nombre"].lower() == nuevo_nombre.lower() for ubic in storage_locations):
        flash("Ya existe una ubicación con ese nombre.", "error")
        return redirect(url_for("gaveta_detalle", nombre=nombre))

    ubicacion["nombre"] = nuevo_nombre
    with get_connection() as conn:
        conn.execute("UPDATE storage_locations SET nombre = ? WHERE lower(nombre) = ?", (nuevo_nombre, nombre.lower()))
        conn.execute("UPDATE inventory_items SET ubicacion = ? WHERE lower(ubicacion) = ?", (nuevo_nombre, nombre.lower()))
        conn.execute(
            "UPDATE gaveta_asignaciones SET gaveta_nombre = ? WHERE lower(gaveta_nombre) = ?",
            (nuevo_nombre, nombre.lower()),
        )

    for articulo in inventory_items:
        if articulo["ubicacion"].lower() == nombre.lower():
            articulo["ubicacion"] = nuevo_nombre

    for asignacion in gaveta_asignaciones.values():
        if asignacion["gaveta"]["nombre"].lower() == nombre.lower():
            asignacion["gaveta"]["nombre"] = nuevo_nombre

    flash("Nombre de gaveta actualizado correctamente.", "success")
    return redirect(url_for("gaveta_detalle", nombre=nuevo_nombre))


@app.route("/crear-gavetas/<path:nombre>/estado", methods=["POST"])
def actualizar_estado_gaveta(nombre: str):
    ubicacion = _buscar_gaveta(nombre)
    if not ubicacion:
        flash("No se encontró la gaveta solicitada.", "error")
        return redirect(url_for("crear_gavetas"))

    nuevo_estado = request.form.get("estado", "").strip().capitalize() or "Abierta"
    if nuevo_estado not in {"Abierta", "Facturada"}:
        flash("Selecciona un estado válido para la gaveta.", "error")
        return redirect(url_for("gaveta_detalle", nombre=nombre))

    if nuevo_estado == ubicacion.get("estado", "Abierta"):
        flash("La gaveta ya tiene el estado seleccionado.", "info")
        return redirect(url_for("gaveta_detalle", nombre=nombre))

    if nuevo_estado == "Facturada":
        _archivar_asignaciones_gaveta(ubicacion["nombre"])
        ubicacion["estado"] = nuevo_estado
        with get_connection() as conn:
            conn.execute(
                "UPDATE storage_locations SET estado = ? WHERE lower(nombre) = lower(?)",
                (ubicacion["estado"], ubicacion["nombre"]),
            )
        nueva_gaveta = _crear_gaveta_sucesora(ubicacion["nombre"], ubicacion["tipo"])
        flash(
            f"La gaveta se marcó como facturada. Se creó {nueva_gaveta['nombre']} para nuevas unidades.",
            "info",
        )
        return redirect(url_for("gaveta_detalle", nombre=ubicacion["nombre"]))

    ubicacion["estado"] = nuevo_estado
    with get_connection() as conn:
        conn.execute(
            "UPDATE storage_locations SET estado = ? WHERE lower(nombre) = lower(?)",
            (ubicacion["estado"], ubicacion["nombre"]),
        )

    flash("La gaveta volvió a estar abierta.", "success")
    return redirect(url_for("gaveta_detalle", nombre=ubicacion["nombre"]))


@app.route("/crear-gavetas/<path:nombre>/eliminar", methods=["POST"])
def eliminar_gaveta(nombre: str):
    global storage_locations, gaveta_asignaciones

    ubicacion = next(
        (
            ubicacion
            for ubicacion in storage_locations
            if ubicacion["nombre"].lower() == nombre.lower()
        ),
        None,
    )
    if not ubicacion:
        flash("No se encontró la ubicación a eliminar.", "error")
        return redirect(url_for("crear_gavetas"))

    asignaciones_a_eliminar = [
        clave
        for clave, asignacion in gaveta_asignaciones.items()
        if asignacion["gaveta"]["nombre"].lower() == ubicacion["nombre"].lower()
    ]
    for clave in asignaciones_a_eliminar:
        del gaveta_asignaciones[clave]

    storage_locations = [
        ubic
        for ubic in storage_locations
        if ubic["nombre"].lower() != ubicacion["nombre"].lower()
    ]

    with get_connection() as conn:
        conn.execute(
            "DELETE FROM gaveta_asignaciones WHERE lower(gaveta_nombre) = lower(?)",
            (ubicacion["nombre"],),
        )
        conn.execute(
            "DELETE FROM storage_locations WHERE lower(nombre) = lower(?)",
            (ubicacion["nombre"],),
        )

    flash("La ubicación se eliminó correctamente.", "success")
    return redirect(url_for("crear_gavetas"))


@app.route("/crear-gavetas/<path:nombre>")
def gaveta_detalle(nombre: str):
    ubicacion = next(
        (
            ubicacion
            for ubicacion in storage_locations
            if ubicacion["nombre"].lower() == nombre.lower()
        ),
        None,
    )
    if not ubicacion:
        flash("No se encontró la ubicación solicitada.", "error")
        return redirect(url_for("crear_gavetas"))

    articulos = [
        item
        for item in inventory_items
        if item["ubicacion"].lower() == ubicacion["nombre"].lower()
    ]
    asignaciones_gaveta = [
        asignacion
        for asignacion in gaveta_asignaciones.values()
        if asignacion["gaveta"]["nombre"].lower() == ubicacion["nombre"].lower()
    ]
    asignaciones_gaveta_archivadas = [
        asignacion
        for asignacion in gaveta_asignaciones_archivadas
        if asignacion["gaveta"]["nombre"].lower() == ubicacion["nombre"].lower()
    ]
    asignaciones_gaveta.sort(key=lambda asignacion: (asignacion["pedido_id"], asignacion["codigo"].lower()))
    asignaciones_gaveta_archivadas.sort(
        key=lambda asignacion: (asignacion["pedido_id"], asignacion["codigo"].lower())
    )
    total_unidades_inventario = sum(item["cantidad"] for item in articulos)
    total_unidades_asignadas = sum(asignacion["unidades"] for asignacion in asignaciones_gaveta)
    total_unidades = total_unidades_inventario + total_unidades_asignadas

    return render_template(
        "gaveta_detalle.html",
        ubicacion=ubicacion,
        articulos=articulos,
        total_unidades=total_unidades,
        total_unidades_inventario=total_unidades_inventario,
        total_unidades_asignadas=total_unidades_asignadas,
        asignaciones_gaveta=asignaciones_gaveta,
        asignaciones_gaveta_archivadas=asignaciones_gaveta_archivadas,
    )


def _lineas_pendientes():
    lineas = []
    for pedido in purchase_orders:
        if pedido.get("cerrado"):
            continue
        for linea in pedido["lineas"]:
            if linea["cantidad_pendiente"] > 0:
                lineas.append(
                    {
                        "pedido_id": pedido["id"],
                        "cliente": pedido["cliente"],
                        "codigo": linea["codigo"],
                        "descripcion": linea["descripcion"],
                        "cantidad_pedida": linea["cantidad_pedida"],
                        "cantidad_recibida": linea["cantidad_recibida"],
                        "cantidad_pendiente": linea["cantidad_pendiente"],
                        "fecha": pedido["fecha"],
                    }
                )
    return lineas


def _buscar_linea_en_pedido(pedido_id: int, codigo: str):
    pedido = next((p for p in purchase_orders if p["id"] == pedido_id), None)
    if not pedido:
        return None, None
    linea = next(
        (l for l in pedido["lineas"] if l["codigo"].lower() == codigo.lower()), None
    )
    return pedido, linea


def _clave_gaveta(pedido_id: int, codigo: str):
    return (pedido_id, codigo.lower())


def _generar_nombre_gaveta() -> str:
    global gaveta_secuencia
    nombre = f"Gaveta #{gaveta_secuencia}"
    gaveta_secuencia += 1
    return nombre


def _buscar_gaveta(nombre: str):
    return next(
        (gaveta for gaveta in storage_locations if gaveta["nombre"].lower() == nombre.lower()),
        None,
    )


def _es_gaveta_facturada(gaveta_nombre: str) -> bool:
    ubicacion = _buscar_gaveta(gaveta_nombre)
    estado = (ubicacion or {}).get("estado", "Abierta")
    return estado.lower() == "facturada"


def _crear_gaveta_sucesora(nombre_base: str, tipo: str):
    existente = {gaveta["nombre"].lower() for gaveta in storage_locations}
    indice = 2
    nuevo_nombre = f"{nombre_base} ({indice})"
    while nuevo_nombre.lower() in existente:
        indice += 1
        nuevo_nombre = f"{nombre_base} ({indice})"

    nueva_gaveta = {"nombre": nuevo_nombre, "tipo": tipo, "estado": "Abierta", "created_at": datetime.now()}
    storage_locations.append(nueva_gaveta)
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO storage_locations (nombre, tipo, created_at, estado) VALUES (?, ?, ?, ?)",
            (nueva_gaveta["nombre"], nueva_gaveta["tipo"], nueva_gaveta["created_at"].isoformat(), nueva_gaveta["estado"]),
        )
    return nueva_gaveta


def _obtener_o_crear_gaveta(pedido: dict, linea: dict):
    clave = _clave_gaveta(pedido["id"], linea["codigo"])
    asignacion = gaveta_asignaciones.get(clave)
    if asignacion and _es_gaveta_facturada(asignacion["gaveta"]["nombre"]):
        _archivar_asignaciones_gaveta(asignacion["gaveta"]["nombre"])
        asignacion = None

    if asignacion:
        return clave, asignacion, False

    asignacion_existente = next(
        (asig for asig in gaveta_asignaciones.values() if asig["pedido_id"] == pedido["id"]),
        None,
    )

    if asignacion_existente:
        gaveta = asignacion_existente["gaveta"]
        gaveta_creada = False
    else:
        gaveta_existente = next(
            (
                ubicacion
                for ubicacion in storage_locations
                if ubicacion["tipo"].lower() == "gaveta"
                and ubicacion["nombre"].lower() == pedido["nombre"].lower()
            ),
            None,
        )

        if gaveta_existente:
            gaveta = gaveta_existente
            gaveta_creada = False
        else:
            gaveta = _asegurar_gaveta_existente(pedido["nombre"])
            gaveta_creada = True

    fecha_creacion_gaveta = (
        asignacion_existente["gaveta"]["created_at"] if asignacion_existente else gaveta["created_at"]
    )

    asignacion = {
        "pedido_id": pedido["id"],
        "cliente": pedido["cliente"],
        "codigo": linea["codigo"],
        "descripcion": linea.get("descripcion") or linea.get("nombre", linea["codigo"]),
        "unidades": 0,
        "gaveta": gaveta,
    }
    gaveta_asignaciones[clave] = asignacion
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO gaveta_asignaciones (pedido_id, codigo, cliente, descripcion, unidades, gaveta_nombre, gaveta_tipo, created_at, activa)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (
                asignacion["pedido_id"],
                asignacion["codigo"],
                asignacion["cliente"],
                asignacion["descripcion"],
                asignacion["unidades"],
                gaveta["nombre"],
                gaveta["tipo"],
                fecha_creacion_gaveta.isoformat(),
            ),
        )
    return clave, asignacion, gaveta_creada


def _actualizar_unidades_gaveta(clave, delta: int):
    asignacion = gaveta_asignaciones.get(clave)
    if asignacion:
        asignacion["unidades"] = max(asignacion["unidades"] + delta, 0)
        _ajustar_stock_gaveta(asignacion, delta)
        with get_connection() as conn:
            conn.execute(
                "UPDATE gaveta_asignaciones SET unidades = ? WHERE pedido_id = ? AND lower(codigo) = ?",
                (asignacion["unidades"], clave[0], clave[1].lower()),
            )
    return asignacion


def _listar_gavetas_activas():
    gavetas = {}
    for asignacion in gaveta_asignaciones.values():
        nombre = asignacion["gaveta"]["nombre"]
        if _es_gaveta_facturada(nombre):
            continue
        registro = gavetas.setdefault(nombre, {"nombre": nombre, "unidades": 0})
        registro["unidades"] += asignacion["unidades"]

    return sorted(gavetas.values(), key=lambda gaveta: gaveta["nombre"].lower())


def _asegurar_gaveta_existente(nombre: str):
    existente = next(
        (loc for loc in storage_locations if loc["nombre"].lower() == nombre.lower()),
        None,
    )
    if existente:
        if _es_gaveta_facturada(existente["nombre"]):
            return _crear_gaveta_sucesora(nombre, existente["tipo"])
        return existente

    nueva_gaveta = {"nombre": nombre, "tipo": "Gaveta", "estado": "Abierta", "created_at": datetime.now()}
    storage_locations.append(nueva_gaveta)
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO storage_locations (nombre, tipo, created_at, estado) VALUES (?, ?, ?, ?)",
            (
                nueva_gaveta["nombre"],
                nueva_gaveta["tipo"],
                nueva_gaveta["created_at"].isoformat(),
                nueva_gaveta["estado"],
            ),
        )
    return nueva_gaveta


def _actualizar_destino_gaveta(clave, nuevo_nombre: str):
    asignacion = gaveta_asignaciones.get(clave)
    if not asignacion:
        return None

    nueva_gaveta = _asegurar_gaveta_existente(nuevo_nombre)
    for asignacion_pedido in gaveta_asignaciones.values():
        if asignacion_pedido["pedido_id"] == asignacion["pedido_id"]:
            asignacion_pedido["gaveta"] = nueva_gaveta

    with get_connection() as conn:
        conn.execute(
            "UPDATE gaveta_asignaciones SET gaveta_nombre = ?, gaveta_tipo = ? WHERE pedido_id = ?",
            (nuevo_nombre, nueva_gaveta["tipo"], asignacion["pedido_id"]),
        )
    return asignacion


def _archivar_asignaciones_gaveta(nombre: str):
    claves = [
        clave
        for clave, asignacion in gaveta_asignaciones.items()
        if asignacion["gaveta"]["nombre"].lower() == nombre.lower()
    ]

    if not claves:
        return []

    with get_connection() as conn:
        for clave in claves:
            asignacion = gaveta_asignaciones.pop(clave, None)
            if not asignacion:
                continue
            if asignacion.get("unidades"):
                _ajustar_stock_gaveta(asignacion, -asignacion["unidades"])
            gaveta_asignaciones_archivadas.append(asignacion)
            conn.execute(
                "UPDATE gaveta_asignaciones SET activa = 0 WHERE pedido_id = ? AND lower(codigo) = ?",
                (clave[0], clave[1].lower()),
            )

    return claves


def _ajustar_stock_gaveta(asignacion: dict, delta: int):
    if not asignacion or delta == 0:
        return None

    codigo = asignacion["codigo"]
    ubicacion = asignacion["gaveta"]["nombre"]
    if delta > 0 and _es_gaveta_facturada(ubicacion):
        return None
    descripcion = asignacion.get("descripcion", codigo)
    articulo_referencia = next(
        (item for item in inventory_items if item["codigo"].lower() == codigo.lower()), None
    )
    nombre_articulo = articulo_referencia["nombre"] if articulo_referencia else descripcion
    tipo_articulo = articulo_referencia.get("tipo", "") if articulo_referencia else ""
    precio_pvo = float(articulo_referencia.get("precio_pvo", 0)) if articulo_referencia else 0.0
    precio_pvp = float(articulo_referencia.get("precio_pvp", 0)) if articulo_referencia else 0.0

    existente = next(
        (
            item
            for item in inventory_items
            if item["codigo"].lower() == codigo.lower()
            and item["ubicacion"].lower() == ubicacion.lower()
        ),
        None,
    )

    if existente:
        existente["cantidad"] = max(existente["cantidad"] + delta, 0)
        with get_connection() as conn:
            conn.execute(
                "UPDATE inventory_items SET cantidad = ? WHERE id = ?",
                (existente["cantidad"], existente["id"]),
            )
        return existente

    if delta > 0:
        with get_connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO inventory_items (codigo, nombre, tipo, precio_pvo, precio_pvp, cantidad, ubicacion)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (codigo, nombre_articulo, tipo_articulo, precio_pvo, precio_pvp, delta, ubicacion),
            )
            nuevo_id = cursor.lastrowid
        nuevo_articulo = {
            "id": nuevo_id,
            "codigo": codigo,
            "nombre": nombre_articulo,
            "tipo": tipo_articulo,
            "precio_pvo": precio_pvo,
            "precio_pvp": precio_pvp,
            "cantidad": delta,
            "ubicacion": ubicacion,
        }
        inventory_items.append(nuevo_articulo)
        return nuevo_articulo

    return None


def _transferir_unidades_asignacion(asignacion: dict, nueva_gaveta: dict):
    if not asignacion or asignacion["gaveta"]["nombre"].lower() == nueva_gaveta["nombre"].lower():
        return

    unidades = asignacion.get("unidades", 0)
    if unidades > 0:
        _ajustar_stock_gaveta(asignacion, -unidades)
        asignacion_temporal = {**asignacion, "gaveta": nueva_gaveta}
        _ajustar_stock_gaveta(asignacion_temporal, unidades)
    asignacion["gaveta"] = nueva_gaveta


def _asignar_gaveta_existente(pedido: dict, linea: dict, gaveta: dict):
    clave = _clave_gaveta(pedido["id"], linea["codigo"])
    descripcion = linea.get("descripcion") or linea.get("nombre", linea["codigo"])
    asignacion = gaveta_asignaciones.get(clave)

    if asignacion:
        _transferir_unidades_asignacion(asignacion, gaveta)
        asignacion["descripcion"] = descripcion
        with get_connection() as conn:
            conn.execute(
                """
                UPDATE gaveta_asignaciones
                SET descripcion = ?, gaveta_nombre = ?, gaveta_tipo = ?
                WHERE pedido_id = ? AND lower(codigo) = ?
                """,
                (
                    descripcion,
                    gaveta["nombre"],
                    gaveta["tipo"],
                    pedido["id"],
                    linea["codigo"].lower(),
                ),
            )
        return clave, asignacion, False

    nueva_asignacion = {
        "pedido_id": pedido["id"],
        "cliente": pedido["cliente"],
        "codigo": linea["codigo"],
        "descripcion": descripcion,
        "unidades": 0,
        "gaveta": gaveta,
    }
    gaveta_asignaciones[clave] = nueva_asignacion
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO gaveta_asignaciones (pedido_id, codigo, cliente, descripcion, unidades, gaveta_nombre, gaveta_tipo, created_at, activa)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (
                nueva_asignacion["pedido_id"],
                nueva_asignacion["codigo"],
                nueva_asignacion["cliente"],
                nueva_asignacion["descripcion"],
                nueva_asignacion["unidades"],
                gaveta["nombre"],
                gaveta["tipo"],
                datetime.now().isoformat(),
            ),
        )
    return clave, nueva_asignacion, True


def _totales_albaran(albaran):
    total_unidades = sum(linea["cantidad"] for linea in albaran["lineas"])
    total_pvo = sum(linea["precio_pvo"] * linea["cantidad"] for linea in albaran["lineas"])
    total_pvp = sum(linea["precio_pvp"] * linea["cantidad"] for linea in albaran["lineas"])
    return {
        "total_unidades": total_unidades,
        "total_pvo": total_pvo,
        "total_pvp": total_pvp,
    }


def _buscar_albaran(albaran_id: int):
    return next((nota for nota in delivery_notes if nota["id"] == albaran_id), None)


def _generar_numero_albaran():
    secuencia = max((nota["id"] for nota in delivery_notes), default=7000) - 7000 + 1
    return f"ALB-{datetime.now().year}-{secuencia:03d}"


def _crear_albaran(numero: str | None = None, proveedor: str | None = None, fabrica: str | None = None):
    nuevo_id = max((nota["id"] for nota in delivery_notes), default=7000) + 1
    nuevo_albaran = {
        "id": nuevo_id,
        "numero": numero.strip() if numero and numero.strip() else _generar_numero_albaran(),
        "fecha": datetime.now(),
        "proveedor": proveedor.strip() if proveedor and proveedor.strip() else "Proveedor pendiente",
        "fabrica": fabrica.strip() if fabrica and fabrica.strip() else "Almacén principal",
        "precio_transporte": 0.0,
        "lineas": [],
    }
    delivery_notes.append(nuevo_albaran)
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO delivery_notes (id, numero, fecha, proveedor, fabrica, precio_transporte)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                nuevo_albaran["id"],
                nuevo_albaran["numero"],
                nuevo_albaran["fecha"].isoformat(),
                nuevo_albaran["proveedor"],
                nuevo_albaran["fabrica"],
                nuevo_albaran["precio_transporte"],
            ),
        )
    return nuevo_albaran


def _registrar_en_albaran(albaran: dict, linea: dict):
    if not albaran:
        return None

    codigo = linea.get("codigo", "")
    linea_existente = next(
        (item for item in albaran["lineas"] if item["codigo"].lower() == codigo.lower()),
        None,
    )
    if linea_existente:
        linea_existente["cantidad"] += 1
        with get_connection() as conn:
            conn.execute(
                """
                UPDATE delivery_note_lines
                SET cantidad = ?
                WHERE albaran_id = ? AND lower(codigo) = ?
                """,
                (linea_existente["cantidad"], albaran["id"], codigo.lower()),
            )
        return linea_existente

    nueva_linea = {
        "codigo": codigo,
        "nombre": linea.get("descripcion") or linea.get("nombre", codigo),
        "tipo": linea.get("tipo", ""),
        "precio_pvo": linea.get("precio_pvo", 0.0),
        "precio_pvp": linea.get("precio_pvp", 0.0),
        "cantidad": 1,
    }
    albaran["lineas"].append(nueva_linea)
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO delivery_note_lines (albaran_id, codigo, nombre, tipo, precio_pvo, precio_pvp, cantidad)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                albaran["id"],
                nueva_linea["codigo"],
                nueva_linea["nombre"],
                nueva_linea["tipo"],
                nueva_linea["precio_pvo"],
                nueva_linea["precio_pvp"],
                nueva_linea["cantidad"],
            ),
        )
    return nueva_linea


def _buscar_linea_por_codigo(codigo: str):
    codigo_lower = codigo.lower()
    pedidos_ordenados = sorted(purchase_orders, key=lambda pedido: pedido["fecha"])
    for pedido in pedidos_ordenados:
        if pedido.get("cerrado"):
            continue
        for linea in pedido["lineas"]:
            if linea["codigo"].lower() == codigo_lower and linea["cantidad_pendiente"] > 0:
                return pedido, linea
    return None, None


def _deshacer_ultima_lectura():
    if not lecturas_historial:
        return None

    registro = lecturas_historial.pop()
    linea = registro["linea"]
    if linea["cantidad_recibida"] <= 0:
        return None

    linea["cantidad_recibida"] = max(linea["cantidad_recibida"] - 1, 0)
    linea["cantidad_pendiente"] = min(
        linea["cantidad_pendiente"] + 1, linea["cantidad_pedida"]
    )
    _persistir_linea_pedido(registro["pedido_id"], linea)
    gaveta_key = registro.get("gaveta_key")
    asignacion = _actualizar_unidades_gaveta(gaveta_key, -1) if gaveta_key else None
    if asignacion:
        registro["gaveta"] = asignacion["gaveta"]["nombre"]
        registro["unidades_gaveta"] = asignacion["unidades"]
    return registro


def _paginar(items, pagina, tamano=6):
    total = len(items)
    total_paginas = max(1, math.ceil(total / tamano))
    pagina = max(1, min(pagina, total_paginas))
    inicio = (pagina - 1) * tamano
    return items[inicio : inicio + tamano], total_paginas, total


@app.route("/lectura-codigos", methods=["GET", "POST"])
def lectura_codigos():
    global active_delivery_note_id
    resultado = None
    codigo = ""
    pendiente_cliente = request.args.get("pendiente_cliente", "").strip()
    pendiente_codigo = request.args.get("pendiente_codigo", "").strip()
    pendiente_orden = request.args.get("pendiente_orden", "fecha_desc")
    pendiente_pagina = request.args.get("pendiente_pagina", type=int, default=1)
    gaveta_nombre = request.args.get("gaveta_nombre", "").strip()
    gaveta_orden = request.args.get("gaveta_orden", "nombre")
    gaveta_pagina = request.args.get("gaveta_pagina", type=int, default=1)
    albaran_activo = _buscar_albaran(active_delivery_note_id) if active_delivery_note_id else None

    if request.method == "POST":
        accion = request.form.get("accion")
        if accion == "nuevo_albaran":
            proveedor = request.form.get("proveedor")
            numero_albaran = request.form.get("numero")
            nuevo_albaran = _crear_albaran(numero_albaran, proveedor)
            active_delivery_note_id = nuevo_albaran["id"]
            albaran_activo = nuevo_albaran
            flash(
                f"Albarán {nuevo_albaran['numero']} creado y listo para registrar lecturas.",
                "success",
            )
        elif accion == "seleccionar_albaran":
            albaran_id = request.form.get("albaran_id", type=int)
            seleccionado = _buscar_albaran(albaran_id) if albaran_id else None
            if seleccionado:
                active_delivery_note_id = seleccionado["id"]
                albaran_activo = seleccionado
                flash(
                    f"Leyendo códigos en el albarán {seleccionado['numero']}.", "info"
                )
            else:
                flash("Selecciona un albarán válido para continuar.", "warning")
        elif accion == "detener_albaran":
            active_delivery_note_id = None
            albaran_activo = None
            flash("Se detuvo la lectura en el albarán en curso.", "info")
        elif accion == "deshacer":
            registro = _deshacer_ultima_lectura()
            if registro is None:
                flash("No hay lecturas previas para deshacer.", "warning")
            else:
                flash(
                    f"Se revirtió la última lectura del pedido #{registro['pedido_id']}.",
                    "info",
                )
                resultado = {
                    "pedido_id": registro["pedido_id"],
                    "cliente": registro["cliente"],
                    "linea": registro["linea"],
                    "completado": False,
                    "deshacer": True,
                    "gaveta_creada": False,
                }
                if registro.get("gaveta"):
                    resultado["gaveta"] = registro["gaveta"]
                    resultado["unidades_gaveta"] = registro.get("unidades_gaveta", 0)
                lecturas_registradas.append(
                    {
                        "timestamp": datetime.now(),
                        "pedido_id": registro["pedido_id"],
                        "cliente": registro["cliente"],
                        "codigo": registro["linea"]["codigo"],
                        "descripcion": registro["linea"].get("descripcion", ""),
                        "gaveta": registro.get("gaveta"),
                        "accion": "deshacer",
                    }
                )
        elif accion == "ajustar_linea":
            pedido_id = request.form.get("pedido_id", type=int)
            codigo_linea = request.form.get("codigo_linea", "").strip()
            nuevas_recibidas = request.form.get("nuevas_recibidas", type=int)
            pedido, linea = _buscar_linea_en_pedido(pedido_id, codigo_linea)
            if not pedido or not linea:
                flash("No se encontró la línea a ajustar.", "error")
            else:
                nuevas_recibidas = max(
                    0, min(nuevas_recibidas or 0, linea["cantidad_pedida"])
                )
                linea["cantidad_recibida"] = nuevas_recibidas
                linea["cantidad_pendiente"] = max(
                    linea["cantidad_pedida"] - nuevas_recibidas, 0
                )
                _persistir_linea_pedido(pedido_id, linea)
                flash(
                    f"Actualizada la recepción del código {linea['codigo']} en el pedido #{pedido_id}.",
                    "success",
                )
        elif accion == "actualizar_gaveta_unidades":
            pedido_id = request.form.get("pedido_id", type=int)
            codigo_linea = request.form.get("codigo_linea", "")
            nuevas_unidades = request.form.get("nuevas_unidades", type=int)
            clave = _clave_gaveta(pedido_id, codigo_linea)
            asignacion = gaveta_asignaciones.get(clave)
            if not asignacion:
                flash("No se encontró la gaveta seleccionada.", "error")
            else:
                nuevas_unidades = max(0, nuevas_unidades or 0)
                delta = nuevas_unidades - asignacion["unidades"]
                _actualizar_unidades_gaveta(clave, delta)
                flash("Unidades actualizadas en la gaveta.", "success")
        elif accion == "mover_gaveta":
            pedido_id = request.form.get("pedido_id", type=int)
            codigo_linea = request.form.get("codigo_linea", "")
            nueva_gaveta = request.form.get("nueva_gaveta", "").strip()
            clave = _clave_gaveta(pedido_id, codigo_linea)
            if not nueva_gaveta:
                flash("Introduce un nombre de gaveta válido para mover la asignación.", "error")
            else:
                asignacion = _actualizar_destino_gaveta(clave, nueva_gaveta)
                if asignacion:
                    flash(
                        f"La gaveta del código {codigo_linea} se movió a {nueva_gaveta}.",
                        "info",
                    )
                else:
                    flash("No se pudo mover la gaveta solicitada.", "error")
        else:
            codigo = request.form.get("codigo", "").strip()
            if not codigo:
                flash("Introduce un código de barras.", "error")
            elif not albaran_activo:
                flash(
                    "Crea o selecciona un albarán para registrar las lecturas.", "warning"
                )
            else:
                pedido, linea = _buscar_linea_por_codigo(codigo)
                if not linea:
                    flash("No hay unidades pendientes para ese código.", "warning")
                else:
                    nueva_cantidad_recibida = min(
                        linea["cantidad_recibida"] + 1, linea["cantidad_pedida"]
                    )
                    linea["cantidad_pendiente"] = max(
                        linea["cantidad_pedida"] - nueva_cantidad_recibida, 0
                    )
                    linea["cantidad_recibida"] = nueva_cantidad_recibida
                    _persistir_linea_pedido(pedido["id"], linea)
                    completado = linea["cantidad_pendiente"] == 0
                    gaveta_key, asignacion, gaveta_creada = _obtener_o_crear_gaveta(
                        pedido, linea
                    )
                    _actualizar_unidades_gaveta(gaveta_key, 1)
                    linea_albaran = _registrar_en_albaran(albaran_activo, linea)
                    resultado = {
                        "pedido_id": pedido["id"],
                        "pedido_nombre": pedido.get("nombre", ""),
                        "cliente": pedido["cliente"],
                        "linea": linea,
                        "completado": completado,
                        "gaveta": asignacion["gaveta"]["nombre"],
                        "unidades_gaveta": asignacion["unidades"],
                        "gaveta_creada": gaveta_creada,
                        "albaran": albaran_activo["numero"],
                        "linea_albaran": linea_albaran,
                    }
                    lecturas_historial.append(
                        {
                            "pedido_id": pedido["id"],
                            "cliente": pedido["cliente"],
                            "linea": linea,
                            "gaveta_key": gaveta_key,
                        }
                    )
                    lecturas_registradas.append(
                        {
                            "timestamp": datetime.now(),
                            "pedido_id": pedido["id"],
                            "cliente": pedido["cliente"],
                            "codigo": linea["codigo"],
                            "descripcion": linea.get("descripcion", ""),
                            "gaveta": asignacion["gaveta"]["nombre"],
                            "albaran": albaran_activo["numero"] if albaran_activo else None,
                            "accion": "lectura",
                        }
                    )
                    if completado:
                        flash(
                            f"Se completó la línea del código {linea['codigo']} en el pedido #{pedido['id']}.",
                            "success",
                        )
                    else:
                        flash(
                            f"Registrada 1 unidad para el pedido #{pedido['id']}. Pendientes: {linea['cantidad_pendiente']}.",
                            "success",
                        )
                    codigo = ""

    numero_sugerido = _generar_numero_albaran()
    lineas_pendientes = _lineas_pendientes()
    gavetas_activas = _listar_gavetas_activas()

    def _aplicar_filtros_registros(registros, filtro_cliente, filtro_codigo):
        filtrados = registros
        if filtro_cliente:
            filtrados = [
                reg
                for reg in filtrados
                if filtro_cliente.lower() in reg.get("cliente", "").lower()
            ]
        if filtro_codigo:
            filtrados = [
                reg
                for reg in filtrados
                if filtro_codigo.lower() in reg.get("codigo", "").lower()
                or filtro_codigo.lower() in reg.get("descripcion", "").lower()
            ]
        return filtrados

    lineas_pendientes = _aplicar_filtros_registros(
        lineas_pendientes, pendiente_cliente, pendiente_codigo
    )
    if gaveta_nombre:
        gavetas_activas = [
            gaveta
            for gaveta in gavetas_activas
            if gaveta_nombre.lower() in gaveta["nombre"].lower()
        ]

    ordenes_lineas = {
        "fecha_desc": (lambda l: l["fecha"], True),
        "pendientes": (lambda l: l["cantidad_pendiente"], True),
        "cliente": (lambda l: l["cliente"].lower(), False),
        "codigo": (lambda l: l["codigo"].lower(), False),
    }
    key_lineas, reverse_lineas = ordenes_lineas.get(
        pendiente_orden, (lambda l: l["fecha"], True)
    )
    lineas_pendientes = sorted(lineas_pendientes, key=key_lineas, reverse=reverse_lineas)

    ordenes_gavetas = {
        "nombre": (lambda g: g["nombre"].lower(), False),
        "unidades": (lambda g: g["unidades"], True),
    }
    key_gaveta, reverse_gaveta = ordenes_gavetas.get(
        gaveta_orden, (lambda g: g["nombre"].lower(), False)
    )
    gavetas_activas = sorted(gavetas_activas, key=key_gaveta, reverse=reverse_gaveta)

    lineas_paginadas, paginas_pendientes, total_lineas = _paginar(
        lineas_pendientes, pendiente_pagina
    )
    gavetas_paginadas, paginas_gavetas, total_gavetas = _paginar(
        gavetas_activas, gaveta_pagina
    )

    albaranes_disponibles = sorted(
        delivery_notes, key=lambda nota: nota["fecha"], reverse=True
    )
    return render_template(
        "lectura_codigos.html",
        codigo=codigo,
        resultado=resultado,
        lineas_pendientes=lineas_paginadas,
        gavetas_activas=gavetas_paginadas,
        albaran_activo=albaran_activo,
        albaranes=albaranes_disponibles,
        numero_sugerido=numero_sugerido,
        pendiente_cliente=pendiente_cliente,
        pendiente_codigo=pendiente_codigo,
        pendiente_orden=pendiente_orden,
        pendiente_pagina=pendiente_pagina,
        paginas_pendientes=paginas_pendientes,
        total_lineas=total_lineas,
        gaveta_nombre=gaveta_nombre,
        gaveta_orden=gaveta_orden,
        gaveta_pagina=gaveta_pagina,
        paginas_gavetas=paginas_gavetas,
        total_gavetas=total_gavetas,
        filtros_query=request.args,
    )


@app.route("/historial-lecturas")
def historial_lecturas():
    filtro_cliente = request.args.get("cliente", "").strip()
    filtro_codigo = request.args.get("codigo", "").strip()
    filtro_accion = request.args.get("accion", "").strip().lower()
    filtro_pedido = request.args.get("pedido_id", type=int)
    pagina = request.args.get("pagina", type=int, default=1)

    registros = list(lecturas_registradas)

    if filtro_cliente:
        registros = [
            reg
            for reg in registros
            if filtro_cliente.lower() in reg.get("cliente", "").lower()
        ]

    if filtro_codigo:
        registros = [
            reg
            for reg in registros
            if filtro_codigo.lower() in reg.get("codigo", "").lower()
            or filtro_codigo.lower() in reg.get("descripcion", "").lower()
        ]

    if filtro_accion in {"lectura", "deshacer"}:
        registros = [reg for reg in registros if reg.get("accion") == filtro_accion]

    if filtro_pedido:
        registros = [
            reg for reg in registros if reg.get("pedido_id") == filtro_pedido
        ]

    registros = sorted(
        registros,
        key=lambda reg: reg.get("timestamp") or datetime.min,
        reverse=True,
    )

    registros_paginados, total_paginas, total_registros = _paginar(
        registros, pagina, tamano=15
    )

    return render_template(
        "historial_lecturas.html",
        registros=registros_paginados,
        pagina=pagina,
        total_paginas=total_paginas,
        total_registros=total_registros,
        filtro_cliente=filtro_cliente,
        filtro_codigo=filtro_codigo,
        filtro_accion=filtro_accion,
        filtro_pedido=filtro_pedido,
    )


def _descripcion_por_codigo(codigo: str) -> str:
    articulo = next(
        (item for item in inventory_items if item["codigo"].lower() == codigo.lower()),
        None,
    )
    return articulo["nombre"] if articulo else codigo


def _leer_pedidos_excel(archivo: io.BytesIO):
    try:
        workbook = load_workbook(archivo, data_only=True)
    except Exception as exc:  # pragma: no cover - validación defensiva
        raise ValueError("No se pudo leer el Excel. Verifica el formato.") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")

    headers = [
        str(cell).strip().lower() if cell is not None else "" for cell in rows[0]
    ]
    header_map = {nombre: idx for idx, nombre in enumerate(headers)}
    required_headers = {"pedido", "cliente", "codigo", "cantidad"}

    if not required_headers.issubset(header_map):
        raise ValueError(
            "La plantilla debe incluir las columnas: pedido, cliente, codigo y cantidad."
        )

    def _leer_valor(row, key, default=None):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return default
        valor = row[idx]
        if valor is None:
            return default
        return valor

    resumen = {"procesadas": 0, "omitidas": 0}
    pedidos_encontrados: dict[tuple[str, str], dict] = {}

    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        resumen["procesadas"] += 1
        pedido_nombre = str(_leer_valor(row, "pedido", "")).strip()
        cliente = str(_leer_valor(row, "cliente", "")).strip()
        codigo = str(_leer_valor(row, "codigo", "")).strip()
        cantidad_valor = _leer_valor(row, "cantidad", 0)

        try:
            cantidad = int(float(cantidad_valor))
        except (TypeError, ValueError):
            cantidad = -1

        if not pedido_nombre or not cliente or not codigo or cantidad <= 0:
            resumen["omitidas"] += 1
            continue

        clave_pedido = (pedido_nombre.lower(), cliente.lower())
        pedido = pedidos_encontrados.setdefault(
            clave_pedido, {"nombre": pedido_nombre, "cliente": cliente, "lineas": {}}
        )

        linea = pedido["lineas"].get(codigo.lower())
        if linea:
            linea["cantidad"] += cantidad
        else:
            pedido["lineas"][codigo.lower()] = {"codigo": codigo, "cantidad": cantidad}

    pedidos_list = [
        {"nombre": datos["nombre"], "cliente": datos["cliente"], "lineas": list(datos["lineas"].values())}
        for datos in pedidos_encontrados.values()
    ]

    return pedidos_list, resumen


def _registrar_pedidos_importados(pedidos_desde_excel: list[dict]):
    resultados = {
        "pedidos_creados": 0,
        "lineas_creadas": 0,
        "lineas_actualizadas": 0,
    }

    for pedido_excel in pedidos_desde_excel:
        existente = next(
            (
                pedido
                for pedido in purchase_orders
                if pedido["nombre"].lower() == pedido_excel["nombre"].lower()
                and pedido["cliente"].lower() == pedido_excel["cliente"].lower()
            ),
            None,
        )

        if existente:
            pedido_objetivo = existente
        else:
            nuevo_id = max((pedido["id"] for pedido in purchase_orders), default=0) + 1
            pedido_objetivo = {
                "id": nuevo_id,
                "nombre": pedido_excel["nombre"],
                "cliente": pedido_excel["cliente"],
                "fecha": datetime.now(),
                "estado": "Pendiente",
                "notas": "Importado desde XLSX.",
                "lineas": [],
            }
            purchase_orders.append(pedido_objetivo)
            with get_connection() as conn:
                conn.execute(
                    "INSERT INTO purchase_orders (id, nombre, cliente, fecha, estado, notas) VALUES (?, ?, ?, ?, ?, ?)",
                    (
                        pedido_objetivo["id"],
                        pedido_objetivo["nombre"],
                        pedido_objetivo["cliente"],
                        pedido_objetivo["fecha"].isoformat(),
                        pedido_objetivo["estado"],
                        pedido_objetivo["notas"],
                    ),
                )
            resultados["pedidos_creados"] += 1

        for linea_excel in pedido_excel.get("lineas", []):
            codigo = linea_excel.get("codigo", "")
            cantidad = linea_excel.get("cantidad", 0)
            descripcion = _descripcion_por_codigo(codigo)

            linea_existente = next(
                (
                    linea
                    for linea in pedido_objetivo["lineas"]
                    if linea["codigo"].lower() == codigo.lower()
                ),
                None,
            )

            if linea_existente:
                linea_existente["cantidad_pedida"] += cantidad
                linea_existente["cantidad_pendiente"] += cantidad
                linea_existente["descripcion"] = linea_existente.get("descripcion") or descripcion
                _persistir_linea_pedido(pedido_objetivo["id"], linea_existente)
                resultados["lineas_actualizadas"] += 1
            else:
                nueva_linea = {
                    "codigo": codigo,
                    "descripcion": descripcion,
                    "cantidad_pedida": cantidad,
                    "cantidad_recibida": 0,
                    "cantidad_pendiente": cantidad,
                }
                pedido_objetivo["lineas"].append(nueva_linea)
                _insertar_linea_pedido(pedido_objetivo["id"], nueva_linea)
                resultados["lineas_creadas"] += 1

    return resultados


@app.route("/subir-excel", methods=["GET", "POST"])
def subir_excel():
    resumen = None
    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or archivo.filename == "":
            flash("Selecciona un archivo para subir.", "error")
        elif not archivo.filename.lower().endswith(".xlsx"):
            flash("Formato no soportado. Usa un archivo XLSX.", "error")
        else:
            contenido = archivo.read()
            try:
                pedidos_desde_excel, estadisticas = _leer_pedidos_excel(io.BytesIO(contenido))
            except ValueError as exc:
                flash(str(exc), "error")
            else:
                if not estadisticas["procesadas"]:
                    flash(
                        "No se encontraron filas de datos en el archivo.",
                        "error",
                    )
                else:
                    resultados = _registrar_pedidos_importados(pedidos_desde_excel)
                    resumen = {
                        "nombre": archivo.filename,
                        "tamano_kb": round(len(contenido) / 1024, 2),
                        "procesado": True,
                        "pedidos_detectados": len(pedidos_desde_excel),
                        "filas_procesadas": estadisticas["procesadas"],
                        "filas_omitidas": estadisticas["omitidas"],
                        **resultados,
                    }
                    flash(
                        "Pedidos importados correctamente desde el XLSX.", "success"
                    )

    return render_template("subir_excel.html", resumen=resumen)


@app.route("/subir-excel/plantilla")
def descargar_plantilla_excel():
    headers = ["pedido", "cliente", "codigo", "cantidad"]
    rows: list[list] = []

    output = _crear_excel(headers, rows, "Plantilla pedidos")
    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_pedidos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/subir-excel/pedidos")
def descargar_pedidos_excel():
    headers = ["pedido", "cliente", "codigo", "cantidad"]
    rows: list[list] = []

    for pedido in purchase_orders:
        lineas = pedido.get("lineas") or []
        if not lineas:
            rows.append([pedido["nombre"], pedido["cliente"], "", ""])
            continue

        for linea in lineas:
            rows.append(
                [
                    pedido["nombre"],
                    pedido["cliente"],
                    linea.get("codigo", ""),
                    linea.get("cantidad_pendiente")
                    if linea.get("cantidad_pendiente") is not None
                    else linea.get("cantidad_pedida"),
                ]
            )

    output = _crear_excel(headers, rows, "Pedidos actuales")
    filename = f"pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/exportar-informes")
def exportar_informes():
    total_items = len(inventory_items)
    total_unidades = sum(item["cantidad"] for item in inventory_items)
    return render_template(
        "exportar_informes.html",
        total_items=total_items,
        total_unidades=total_unidades,
    )


@app.route("/exportar-informes/descargar")
def descargar_informe():
    headers = ["Código", "Nombre", "Cantidad", "Ubicación"]
    rows = [
        [item["codigo"], item["nombre"], item["cantidad"], item["ubicacion"]]
        for item in inventory_items
    ]

    output = _crear_excel(headers, rows, "Informe stock")
    filename = f"informe_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _crear_excel(headers: list[str], rows: list[list], sheet_name: str = "Hoja 1") -> io.BytesIO:
    """Genera un archivo Excel en memoria a partir de cabeceras y filas."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _sanitize_sheet_name(sheet_name)[:31] or "Hoja 1"

    worksheet.append(headers)
    for fila in rows:
        worksheet.append(fila)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _sanitize_sheet_name(name: str) -> str:
    """Quita caracteres no válidos en nombres de pestaña de Excel."""

    invalid_chars = "[]:*?/\\"
    return "".join(char for char in name if char not in invalid_chars)


@app.route("/buscar-articulos")
def buscar_articulos():
    termino = request.args.get("q", "").strip()
    resultados = []
    if termino:
        termino_lower = termino.lower()
        resultados = [
            item
            for item in inventory_items
            if termino_lower in item["nombre"].lower()
            or termino_lower in item["codigo"].lower()
        ]
        if resultados:
            flash(f"Se encontraron {len(resultados)} coincidencias.", "success")
        else:
            flash("No se encontraron artículos con ese criterio.", "warning")

    return render_template("buscar_articulos.html", termino=termino, resultados=resultados)


@app.route("/inventario/<codigo>/actualizar", methods=["POST"])
def actualizar_inventario(codigo: str):
    if not _articulos_por_codigo(codigo):
        flash("No se encontró el artículo solicitado.", "error")
        return redirect(url_for("panel_control"))

    flash("Gestiona este artículo desde la vista de detalle.", "info")
    return redirect(url_for("inventario_detalle", codigo=codigo))


@app.route("/inventario/<codigo>")
def inventario_detalle(codigo: str):
    articulos = _articulos_por_codigo(codigo)
    if not articulos:
        flash("No se encontró el artículo solicitado.", "error")
        return redirect(url_for("panel_control"))

    total_unidades = sum(item["cantidad"] for item in articulos)
    ubicaciones = sorted(articulos, key=lambda item: item["ubicacion"].lower())
    articulo_base = articulos[0]
    articulo = {
        "codigo": articulo_base["codigo"],
        "nombre": articulo_base["nombre"],
        "tipo": articulo_base.get("tipo", ""),
        "precio_pvo": float(articulo_base.get("precio_pvo", 0)),
        "precio_pvp": float(articulo_base.get("precio_pvp", 0)),
    }

    return render_template(
        "inventario_detalle.html",
        articulo=articulo,
        ubicaciones=ubicaciones,
        total_unidades=total_unidades,
    )


@app.route("/inventario/<codigo>/editar", methods=["POST"])
def actualizar_articulo(codigo: str):
    articulos = _articulos_por_codigo(codigo)
    if not articulos:
        flash("No se encontró el artículo solicitado.", "error")
        return redirect(url_for("panel_control"))

    nuevo_codigo = request.form.get("codigo", "").strip()
    nuevo_nombre = request.form.get("nombre", "").strip()
    nuevo_tipo = request.form.get("tipo", "").strip()
    nuevo_precio_pvo = request.form.get("precio_pvo", type=float)
    nuevo_precio_pvp = request.form.get("precio_pvp", type=float)
    nueva_cantidad_total = request.form.get("cantidad", type=int)

    if (
        not nuevo_codigo
        or not nuevo_nombre
        or nuevo_precio_pvo is None
        or nuevo_precio_pvp is None
        or nueva_cantidad_total is None
        or nueva_cantidad_total < 0
    ):
        flash("Indica código, nombre, precios y una cantidad total válida.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    if nuevo_precio_pvo < 0 or nuevo_precio_pvp < 0:
        flash("Los precios no pueden ser negativos.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    ubicaciones_actuales = {item["ubicacion"].lower() for item in articulos}
    conflicto = next(
        (
            item
            for item in inventory_items
            if item["codigo"].lower() == nuevo_codigo.lower()
            and item["codigo"].lower() != codigo.lower()
            and item["ubicacion"].lower() in ubicaciones_actuales
        ),
        None,
    )
    if conflicto:
        flash(
            "Ya existe un artículo con ese código en alguna de las ubicaciones actuales.",
            "error",
        )
        return redirect(url_for("inventario_detalle", codigo=codigo))

    total_actual = sum(item["cantidad"] for item in articulos)
    diferencia_total = nueva_cantidad_total - total_actual
    primera_ubicacion = articulos[0]
    nueva_cantidad_principal = primera_ubicacion["cantidad"] + diferencia_total
    if nueva_cantidad_principal < 0:
        flash(
            "No puedes fijar la cantidad total por debajo de las unidades almacenadas en otras ubicaciones.",
            "error",
        )
        return redirect(url_for("inventario_detalle", codigo=codigo))
    primera_ubicacion["cantidad"] = nueva_cantidad_principal

    for item in articulos:
        item["codigo"] = nuevo_codigo
        item["nombre"] = nuevo_nombre
        item["tipo"] = nuevo_tipo
        item["precio_pvo"] = nuevo_precio_pvo
        item["precio_pvp"] = nuevo_precio_pvp

    with get_connection() as conn:
        conn.execute(
            """
            UPDATE inventory_items
            SET codigo = ?, nombre = ?, tipo = ?, precio_pvo = ?, precio_pvp = ?
            WHERE lower(codigo) = ?
            """,
            (nuevo_codigo, nuevo_nombre, nuevo_tipo, nuevo_precio_pvo, nuevo_precio_pvp, codigo.lower()),
        )
        conn.execute(
            "UPDATE inventory_items SET cantidad = ? WHERE id = ?",
            (nueva_cantidad_principal, primera_ubicacion["id"]),
        )

    flash("Datos del artículo actualizados correctamente.", "success")
    return redirect(url_for("inventario_detalle", codigo=nuevo_codigo))


@app.route("/inventario/<codigo>/ubicaciones/<int:item_id>/actualizar", methods=["POST"])
def actualizar_existencia(codigo: str, item_id: int):
    articulos = _articulos_por_codigo(codigo)
    if not articulos:
        flash("No se encontró el artículo solicitado.", "error")
        return redirect(url_for("panel_control"))

    item = next((articulo for articulo in articulos if articulo["id"] == item_id), None)
    if not item:
        flash("No se encontró la ubicación indicada.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    nueva_ubicacion = request.form.get("ubicacion", "").strip()
    nueva_cantidad = request.form.get("cantidad", type=int)

    if not nueva_ubicacion or nueva_cantidad is None or nueva_cantidad < 0:
        flash("Indica una ubicación y una cantidad mayor o igual a 0.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    conflicto = next(
        (
            articulo
            for articulo in inventory_items
            if articulo["codigo"].lower() == codigo.lower()
            and articulo["ubicacion"].lower() == nueva_ubicacion.lower()
            and articulo["id"] != item_id
        ),
        None,
    )
    if conflicto:
        flash("Ya existe stock de este artículo en esa ubicación.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    item["ubicacion"] = nueva_ubicacion
    item["cantidad"] = nueva_cantidad
    item["nombre"] = articulos[0]["nombre"]

    with get_connection() as conn:
        conn.execute(
            "UPDATE inventory_items SET cantidad = ?, ubicacion = ?, nombre = ? WHERE id = ?",
            (nueva_cantidad, nueva_ubicacion, item["nombre"], item_id),
        )

    flash("Stock actualizado para la ubicación seleccionada.", "success")
    return redirect(url_for("inventario_detalle", codigo=codigo))


@app.route("/inventario/<codigo>/ubicaciones/agregar", methods=["POST"])
def agregar_existencia(codigo: str):
    articulos = _articulos_por_codigo(codigo)
    if not articulos:
        flash("No se encontró el artículo solicitado.", "error")
        return redirect(url_for("panel_control"))

    ubicacion = request.form.get("ubicacion", "").strip()
    cantidad = request.form.get("cantidad", type=int)

    if not ubicacion or cantidad is None or cantidad < 0:
        flash("Indica ubicación y cantidad mayor o igual a 0.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    existente = next(
        (
            articulo
            for articulo in inventory_items
            if articulo["codigo"].lower() == codigo.lower()
            and articulo["ubicacion"].lower() == ubicacion.lower()
        ),
        None,
    )
    if existente:
        flash("Ya existe stock de este artículo en esa ubicación.", "error")
        return redirect(url_for("inventario_detalle", codigo=codigo))

    articulo_base = articulos[0]
    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO inventory_items (codigo, nombre, tipo, precio_pvo, precio_pvp, cantidad, ubicacion)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                articulo_base["codigo"],
                articulo_base["nombre"],
                articulo_base.get("tipo", ""),
                float(articulo_base.get("precio_pvo", 0)),
                float(articulo_base.get("precio_pvp", 0)),
                cantidad,
                ubicacion,
            ),
        )
        nuevo_id = cursor.lastrowid

    inventory_items.append(
        {
            "id": nuevo_id,
            "codigo": articulo_base["codigo"],
            "nombre": articulo_base["nombre"],
            "tipo": articulo_base.get("tipo", ""),
            "precio_pvo": float(articulo_base.get("precio_pvo", 0)),
            "precio_pvp": float(articulo_base.get("precio_pvp", 0)),
            "cantidad": cantidad,
            "ubicacion": ubicacion,
        }
    )

    flash("Ubicación añadida al artículo.", "success")
    return redirect(url_for("inventario_detalle", codigo=codigo))


@app.route("/inventario/<codigo>/eliminar", methods=["POST"])
def eliminar_articulo(codigo: str):
    articulos = _articulos_por_codigo(codigo)
    if not articulos:
        flash("No se encontró el artículo solicitado.", "error")
        return redirect(url_for("panel_control"))

    inventario_filtrado = [
        item for item in inventory_items if item["codigo"].lower() != codigo.lower()
    ]
    inventory_items[:] = inventario_filtrado

    claves_asignaciones = [
        clave
        for clave, asignacion in gaveta_asignaciones.items()
        if asignacion["codigo"].lower() == codigo.lower()
    ]
    for clave in claves_asignaciones:
        del gaveta_asignaciones[clave]

    with get_connection() as conn:
        conn.execute("DELETE FROM inventory_items WHERE lower(codigo) = ?", (codigo.lower(),))
        conn.execute("DELETE FROM gaveta_asignaciones WHERE lower(codigo) = ?", (codigo.lower(),))

    flash("Artículo eliminado junto con sus ubicaciones.", "success")
    return redirect(url_for("panel_control"))


@app.route("/panel-control")
def panel_control():
    total_articulos = len(inventory_items)
    total_unidades = sum(item["cantidad"] for item in inventory_items)
    ubicaciones_registradas = len(storage_locations)
    bajo_stock = [item for item in inventory_items if item["cantidad"] < 20]

    ubicaciones_recientes = sorted(
        storage_locations, key=lambda ubicacion: ubicacion["created_at"], reverse=True
    )[:5]

    return render_template(
        "panel_control.html",
        total_articulos=total_articulos,
        total_unidades=total_unidades,
        ubicaciones_registradas=ubicaciones_registradas,
        bajo_stock=bajo_stock,
        ubicaciones_recientes=ubicaciones_recientes,
    )


@app.route("/pedidos", methods=["GET", "POST"])
def pedidos():
    if request.method == "POST":
        nombre_pedido = request.form.get("nombre", "").strip()
        cliente = request.form.get("cliente", "").strip()
        codigo = request.form.get("codigo", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        cantidad = request.form.get("cantidad", type=int)

        if not cliente or not codigo or not descripcion or cantidad is None or cantidad <= 0:
            flash("Completa todos los datos del pedido con cantidades válidas.", "error")
        else:
            nuevo_id = max((pedido["id"] for pedido in purchase_orders), default=5000) + 1
            nombre_pedido = nombre_pedido or f"Pedido #{nuevo_id}"
            nueva_linea = {
                "codigo": codigo,
                "descripcion": descripcion,
                "cantidad_pedida": cantidad,
                "cantidad_recibida": 0,
                "cantidad_pendiente": cantidad,
            }
            nuevo_pedido = {
                "id": nuevo_id,
                "nombre": nombre_pedido,
                "cliente": cliente,
                "fecha": datetime.now(),
                "estado": "Pendiente",
                "cerrado": False,
                "notas": "Creado manualmente desde la pantalla de pedidos.",
                "lineas": [nueva_linea],
            }
            purchase_orders.append(nuevo_pedido)
            with get_connection() as conn:
                conn.execute(
                    "INSERT INTO purchase_orders (id, nombre, cliente, fecha, estado, notas, cerrado) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (
                        nuevo_pedido["id"],
                        nuevo_pedido["nombre"],
                        nuevo_pedido["cliente"],
                        nuevo_pedido["fecha"].isoformat(),
                        nuevo_pedido["estado"],
                        nuevo_pedido["notas"],
                        int(nuevo_pedido["cerrado"]),
                    ),
                )
            _insertar_linea_pedido(nuevo_id, nueva_linea)
            flash(f"Pedido #{nuevo_id} registrado correctamente.", "success")
        return redirect(url_for("pedidos"))

    total_lineas = sum(len(pedido["lineas"]) for pedido in purchase_orders)
    total_unidades_pedidas = sum(
        sum(linea["cantidad_pedida"] for linea in pedido["lineas"])
        for pedido in purchase_orders
    )
    total_unidades_pendientes = sum(
        sum(linea["cantidad_pendiente"] for linea in pedido["lineas"])
        for pedido in purchase_orders
        if not pedido.get("cerrado")
    )
    pedidos_abiertos = sum(
        1
        for pedido in purchase_orders
        if not pedido.get("cerrado")
        and any(linea["cantidad_pendiente"] > 0 for linea in pedido["lineas"])
    )

    return render_template(
        "pedidos.html",
        pedidos=purchase_orders,
        total_lineas=total_lineas,
        total_unidades_pedidas=total_unidades_pedidas,
        total_unidades_pendientes=total_unidades_pendientes,
        pedidos_abiertos=pedidos_abiertos,
    )


@app.route("/pedidos/<int:pedido_id>/estado-cierre", methods=["POST"])
def actualizar_estado_cierre_pedido(pedido_id: int):
    pedido = next((pedido for pedido in purchase_orders if pedido["id"] == pedido_id), None)
    if not pedido:
        flash("No se encontró el pedido solicitado.", "error")
        return redirect(url_for("pedidos"))

    cerrado = request.form.get("cerrado", "0") == "1"
    pedido["cerrado"] = cerrado

    with get_connection() as conn:
        conn.execute(
            "UPDATE purchase_orders SET cerrado = ? WHERE id = ?",
            (int(cerrado), pedido_id),
        )

    mensaje = "Pedido cerrado para nuevas lecturas." if cerrado else "Pedido reabierto."
    flash(mensaje, "info")
    return redirect(url_for("pedido_detalle", pedido_id=pedido_id))


@app.route("/pedidos/<int:pedido_id>/eliminar", methods=["POST"])
def eliminar_pedido(pedido_id: int):
    global purchase_orders, gaveta_asignaciones

    pedido = next((pedido for pedido in purchase_orders if pedido["id"] == pedido_id), None)
    if not pedido:
        flash("No se encontró el pedido a eliminar.", "error")
        return redirect(url_for("pedidos"))

    purchase_orders = [pedido for pedido in purchase_orders if pedido["id"] != pedido_id]
    gaveta_asignaciones = {
        clave: asignacion
        for clave, asignacion in gaveta_asignaciones.items()
        if asignacion["pedido_id"] != pedido_id
    }
    global gaveta_asignaciones_archivadas
    gaveta_asignaciones_archivadas = [
        asignacion
        for asignacion in gaveta_asignaciones_archivadas
        if asignacion["pedido_id"] != pedido_id
    ]

    with get_connection() as conn:
        conn.execute("DELETE FROM gaveta_asignaciones WHERE pedido_id = ?", (pedido_id,))
        conn.execute("DELETE FROM purchase_orders WHERE id = ?", (pedido_id,))

    flash("El pedido se eliminó correctamente.", "success")
    return redirect(url_for("pedidos"))


@app.route("/pedidos/<int:pedido_id>/editar", methods=["POST"])
def editar_pedido(pedido_id: int):
    pedido = next((pedido for pedido in purchase_orders if pedido["id"] == pedido_id), None)
    if not pedido:
        flash("No se encontró el pedido a actualizar.", "error")
        return redirect(url_for("pedidos"))

    cliente = request.form.get("cliente", "").strip()
    nombre = request.form.get("nombre", "").strip()
    estado = request.form.get("estado", "").strip()
    notas = request.form.get("notas", "").strip()
    fecha_str = request.form.get("fecha", "").strip()

    if not cliente or not nombre or not estado or not fecha_str:
        flash("Completa nombre, cliente, estado y fecha para actualizar el pedido.", "error")
        return redirect(url_for("pedido_detalle", pedido_id=pedido_id))

    try:
        fecha = datetime.fromisoformat(fecha_str)
    except ValueError:
        flash("La fecha indicada no es válida.", "error")
        return redirect(url_for("pedido_detalle", pedido_id=pedido_id))

    pedido.update(
        {
            "nombre": nombre,
            "cliente": cliente,
            "estado": estado,
            "fecha": fecha,
            "notas": notas,
        }
    )

    with get_connection() as conn:
        conn.execute(
            "UPDATE purchase_orders SET nombre = ?, cliente = ?, fecha = ?, estado = ?, notas = ? WHERE id = ?",
            (nombre, cliente, fecha.isoformat(), estado, notas, pedido_id),
        )

    flash("Pedido actualizado correctamente.", "success")
    return redirect(url_for("pedido_detalle", pedido_id=pedido_id))


@app.route("/pedidos/<int:pedido_id>/lineas", methods=["POST"])
def agregar_linea_pedido(pedido_id: int):
    pedido = next((pedido for pedido in purchase_orders if pedido["id"] == pedido_id), None)
    if not pedido:
        flash("No se encontró el pedido especificado.", "error")
        return redirect(url_for("pedidos"))

    codigo = request.form.get("codigo", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    cantidad = request.form.get("cantidad", type=int)

    if not codigo or not descripcion or cantidad is None or cantidad <= 0:
        flash("Indica código, descripción y una cantidad mayor que cero.", "error")
        return redirect(url_for("pedido_detalle", pedido_id=pedido_id))

    nueva_linea = {
        "codigo": codigo,
        "descripcion": descripcion,
        "cantidad_pedida": cantidad,
        "cantidad_recibida": 0,
        "cantidad_pendiente": cantidad,
    }

    pedido["lineas"].append(nueva_linea)

    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO purchase_order_lines (pedido_id, codigo, descripcion, cantidad_pedida, cantidad_recibida, cantidad_pendiente)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (pedido_id, codigo, descripcion, cantidad, 0, cantidad),
        )

    flash("Se añadió una nueva línea al pedido.", "success")
    return redirect(url_for("pedido_detalle", pedido_id=pedido_id))


@app.route("/pedidos/<int:pedido_id>/asignar-gaveta", methods=["POST"])
def asignar_gaveta_pedido(pedido_id: int):
    pedido = next((pedido for pedido in purchase_orders if pedido["id"] == pedido_id), None)
    if not pedido:
        flash("No se encontró el pedido solicitado.", "error")
        return redirect(url_for("pedidos"))

    codigo = request.form.get("codigo", "").strip()
    gaveta_nombre = request.form.get("gaveta_nombre", "").strip()

    if not codigo or not gaveta_nombre:
        flash("Selecciona un código y una gaveta válida.", "warning")
        return redirect(url_for("pedido_detalle", pedido_id=pedido_id))

    linea = next(
        (linea for linea in pedido["lineas"] if linea["codigo"].lower() == codigo.lower()),
        None,
    )
    if not linea:
        flash("No se encontró la línea indicada en el pedido.", "error")
        return redirect(url_for("pedido_detalle", pedido_id=pedido_id))

    gaveta = next(
        (
            ubicacion
            for ubicacion in storage_locations
            if ubicacion["tipo"].lower() == "gaveta"
            and ubicacion["nombre"].lower() == gaveta_nombre.lower()
        ),
        None,
    )
    if not gaveta:
        flash("Debes elegir una gaveta existente.", "error")
        return redirect(url_for("pedido_detalle", pedido_id=pedido_id))

    _asignar_gaveta_existente(pedido, linea, gaveta)
    flash(
        f"Se asignó la gaveta {gaveta['nombre']} al código {linea['codigo']} del pedido #{pedido_id}.",
        "success",
    )
    return redirect(url_for("pedido_detalle", pedido_id=pedido_id))


@app.route("/pedidos/<int:pedido_id>")
def pedido_detalle(pedido_id: int):
    pedido = next((pedido for pedido in purchase_orders if pedido["id"] == pedido_id), None)
    if not pedido:
        flash("No se encontró el pedido solicitado.", "error")
        return redirect(url_for("pedidos"))

    total_solicitado = sum(linea["cantidad_pedida"] for linea in pedido["lineas"])
    total_recibido = sum(linea["cantidad_recibida"] for linea in pedido["lineas"])
    total_pendiente = sum(linea["cantidad_pendiente"] for linea in pedido["lineas"])

    asignaciones = {
        clave: asignacion
        for clave, asignacion in gaveta_asignaciones.items()
        if asignacion["pedido_id"] == pedido_id
    }
    gavetas_existentes = [
        ubicacion for ubicacion in storage_locations if ubicacion["tipo"].lower() == "gaveta"
    ]

    return render_template(
        "pedido_detalle.html",
        pedido=pedido,
        total_solicitado=total_solicitado,
        total_recibido=total_recibido,
        total_pendiente=total_pendiente,
        asignaciones=asignaciones,
        gavetas_existentes=gavetas_existentes,
    )


@app.route("/albaranes")
def albaranes():
    albaranes_ordenados = sorted(delivery_notes, key=lambda albaran: albaran["fecha"], reverse=True)
    albaranes_enriquecidos = []
    for albaran in albaranes_ordenados:
        totales = _totales_albaran(albaran)
        albaranes_enriquecidos.append({**albaran, **totales})

    totales_generales = {
        "total_albaranes": len(delivery_notes),
        "unidades_recibidas": sum(_totales_albaran(albaran)["total_unidades"] for albaran in delivery_notes),
        "valor_pvo": sum(_totales_albaran(albaran)["total_pvo"] for albaran in delivery_notes),
        "valor_pvp": sum(_totales_albaran(albaran)["total_pvp"] for albaran in delivery_notes),
    }

    return render_template(
        "albaranes.html",
        albaranes=albaranes_enriquecidos,
        totales_generales=totales_generales,
    )


@app.route("/albaranes/<int:albaran_id>/eliminar", methods=["POST"])
def eliminar_albaran(albaran_id: int):
    global delivery_notes

    albaran = next((nota for nota in delivery_notes if nota["id"] == albaran_id), None)
    if not albaran:
        flash("No se encontró el albarán a eliminar.", "error")
        return redirect(url_for("albaranes"))

    delivery_notes = [nota for nota in delivery_notes if nota["id"] != albaran_id]

    with get_connection() as conn:
        conn.execute("DELETE FROM delivery_notes WHERE id = ?", (albaran_id,))

    flash("El albarán se eliminó correctamente.", "success")
    return redirect(url_for("albaranes"))


@app.route("/albaranes/<int:albaran_id>")
def albaran_detalle(albaran_id: int):
    albaran = next((nota for nota in delivery_notes if nota["id"] == albaran_id), None)
    if not albaran:
        flash("No se encontró el albarán solicitado.", "error")
        return redirect(url_for("albaranes"))

    totales = _totales_albaran(albaran)

    return render_template(
        "albaran_detalle.html",
        albaran=albaran,
        totales=totales,
    )


@app.route("/albaranes/<int:albaran_id>/actualizar", methods=["POST"])
def actualizar_albaran(albaran_id: int):
    albaran = next((nota for nota in delivery_notes if nota["id"] == albaran_id), None)
    if not albaran:
        flash("No se encontró el albarán solicitado.", "error")
        return redirect(url_for("albaranes"))

    nuevo_numero = request.form.get("numero", "").strip()
    nuevo_proveedor = request.form.get("proveedor", "").strip()

    if not nuevo_numero or not nuevo_proveedor:
        flash("Completa el número de albarán y el proveedor para actualizar.", "error")
        return redirect(url_for("albaran_detalle", albaran_id=albaran_id))

    albaran["numero"] = nuevo_numero
    albaran["proveedor"] = nuevo_proveedor

    with get_connection() as conn:
        conn.execute(
            "UPDATE delivery_notes SET numero = ?, proveedor = ? WHERE id = ?",
            (nuevo_numero, nuevo_proveedor, albaran_id),
        )

    flash("Datos de cabecera del albarán actualizados.", "success")
    return redirect(url_for("albaran_detalle", albaran_id=albaran_id))


if __name__ == "__main__":
    app.run(debug=True)